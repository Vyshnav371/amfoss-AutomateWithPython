import openpyxl
file=input("enter spreadsheet name")
wb = openpyxl.load_workbook(file)
sheet=wb.active
a,b=sheet.max_row,sheet.max_column
data=[]
for i in range(1,b+1):
    l=[]
    for j in range(1,a+1):
        l.append(sheet.cell(row=j,column=i).value)
    data.append(l)
for i in range(len(data)):
    filename='newfile'+str(i+1)+'.txt'
    file=open(filename,'a')
    for j in range(len(data[i])):
        if data[i][j]!= None:
            file.write(data[i][j]+'\n')
    file.close()

     
