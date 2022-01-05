import sys,openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
data=[]
n=int(input("Enter number of text files: "))
for i in range(n):
    a=input("Enter file name")
    if '.txt' not in a:
        a=a+'.txt'
    s=open(a,'r')
    r=s.readlines()
    for i in range(len(r)):
        r[i]=r[i][:-1]
    data.append(r)
    s.close()
wb = openpyxl.Workbook()
sheet=wb['Sheet']
for i in range(1,len(data)+1):
    for j in range(1,len(data[i-1])+1):
        sheet.cell(row=j,column=i).value=data[i-1][j-1]
wb.save('text-spreadsheet.xlsx')
    
        
