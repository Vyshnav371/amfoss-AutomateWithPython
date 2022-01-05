import sys,openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
n=int(sys.argv[1])
wb = openpyxl.Workbook()
sheet=wb['Sheet']
for i in range(2,n+2):
    char=get_column_letter(i)
    sheet['A'+str(i)]=i-1
    sheet[char+'1']=i-1
for i in range(2,n+2):
    char=get_column_letter(i)
    for j in range(2,n+2):
        sheet[char+str(j)]=(sheet['A'+str(j)].value)*(sheet[char+'1'].value)
wb.save('multiplicationtable.xlsx')
        
        
    
    
        
        
    
