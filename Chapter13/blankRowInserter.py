import sys,openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
n=int(sys.argv[1])
m=int(sys.argv[2])
file=sys.argv[3]
wb = openpyxl.load_workbook(file)
sheet=wb.active
a,b=sheet.max_row,sheet.max_column
data=[]
for i in range(1,a+1):
    l=[]
    for j in range(1,b+1):
        l.append(sheet.cell(row=i, column=j).value)
    data.append(l)
wb = openpyxl.Workbook()
sheet=wb['Sheet']
for i in range(1,n):
    for j in range(1,b+1):
        sheet.cell(row=i, column=j).value=data[i-1][j-1]
for i in range(n+m,b+m+1):
    for j in range(1,b+1):
        sheet.cell(row=i, column=j).value=data[i-m-1][j-1]
wb.save("BlankRow.xlsx")
        
        
